from fastapi import APIRouter, HTTPException, Query
from datetime import datetime, date, timedelta
from typing import Optional
from fastapi.responses import StreamingResponse

from backend.database.mongodb import get_notificaciones_collection, get_ruta_asignada_collection
from backend.database.mongodb import get_transacciones_collection
import io
import pandas as pd
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

router = APIRouter(prefix="/api/reportes", tags=["Reportes"])

@router.get("/transacciones/count")
async def contar_transacciones_por_periodo(
    fecha_inicio: str = Query(..., description="Fecha de inicio en formato YYYY-MM-DD"),
    fecha_fin: str = Query(..., description="Fecha de fin en formato YYYY-MM-DD")
):
    """
    Cuenta transacciones de pago_viaje en un periodo específico
    """
    try:
        collection = get_transacciones_collection()
        
        # Convertir strings a datetime
        inicio = datetime.strptime(fecha_inicio, "%Y-%m-%d")
        fin = datetime.strptime(fecha_fin, "%Y-%m-%d") + timedelta(days=1) - timedelta(seconds=1)
        
        # Validar que la fecha de inicio no sea mayor a la de fin
        if inicio > fin:
            raise HTTPException(status_code=400, detail="La fecha de inicio no puede ser mayor a la fecha de fin")
        
        count = await collection.count_documents({
            "tipo": "pago_viaje",
            "fecha_transaccion": {
                "$gte": inicio,
                "$lte": fin
            }
        })
        
        return {
            "success": True,
            "count": count,
            "fecha_inicio": fecha_inicio,
            "fecha_fin": fecha_fin,
            "tipo": "pago_viaje"
        }
        
    except ValueError as e:
        raise HTTPException(status_code=400, detail="Formato de fecha inválido. Use YYYY-MM-DD")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error al contar transacciones: {str(e)}")

@router.get("/transacciones/suma-total")
async def sumar_monto_transacciones_por_periodo(
    fecha_inicio: str = Query(..., description="Fecha de inicio en formato YYYY-MM-DD"),
    fecha_fin: str = Query(..., description="Fecha de fin en formato YYYY-MM-DD")
):
    """
    Suma el monto total de transacciones de pago_viaje en un periodo específico
    """
    try:
        collection = get_transacciones_collection()
        
        # Convertir strings a datetime
        inicio = datetime.strptime(fecha_inicio, "%Y-%m-%d")
        fin = datetime.strptime(fecha_fin, "%Y-%m-%d") + timedelta(days=1) - timedelta(seconds=1)
        
        # Validar que la fecha de inicio no sea mayor a la de fin
        if inicio > fin:
            raise HTTPException(status_code=400, detail="La fecha de inicio no puede ser mayor a la fecha de fin")
        
        pipeline = [
            {
                "$match": {
                    "tipo": "pago_viaje",
                    "fecha_transaccion": {
                        "$gte": inicio,
                        "$lte": fin
                    }
                }
            },
            {
                "$group": {
                    "_id": None,
                    "total_recaudado": {"$sum": "$monto"},
                    "cantidad_transacciones": {"$sum": 1}
                }
            }
        ]
        
        cursor = collection.aggregate(pipeline)
        result = await cursor.to_list(length=1)
        
        total_recaudado = result[0]["total_recaudado"] if result else 0
        cantidad_transacciones = result[0]["cantidad_transacciones"] if result else 0
        
        return {
            "success": True,
            "total_recaudado": total_recaudado,
            "total_recaudado_formateado": f"${total_recaudado:,.2f}",
            "cantidad_transacciones": cantidad_transacciones,
            "fecha_inicio": fecha_inicio,
            "fecha_fin": fecha_fin,
            "tipo": "pago_viaje"
        }
        
    except ValueError as e:
        raise HTTPException(status_code=400, detail="Formato de fecha inválido. Use YYYY-MM-DD")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error al sumar transacciones: {str(e)}")

@router.get("/transacciones/suma-por-dia")
async def sumar_monto_transacciones_por_dia(
    fecha_inicio: str = Query(..., description="Fecha de inicio en formato YYYY-MM-DD"),
    fecha_fin: str = Query(..., description="Fecha de fin en formato YYYY-MM-DD")
):
    """
    Suma el monto de transacciones de pago_viaje agrupado por día
    """
    try:
        collection = get_transacciones_collection()
        
        # Convertir strings a datetime
        inicio = datetime.strptime(fecha_inicio, "%Y-%m-%d")
        fin = datetime.strptime(fecha_fin, "%Y-%m-%d") + timedelta(days=1) - timedelta(seconds=1)
        
        # Validar que la fecha de inicio no sea mayor a la de fin
        if inicio > fin:
            raise HTTPException(status_code=400, detail="La fecha de inicio no puede ser mayor a la fecha de fin")
        
        pipeline = [
            {
                "$match": {
                    "tipo": "pago_viaje",
                    "fecha_transaccion": {
                        "$gte": inicio,
                        "$lte": fin
                    }
                }
            },
            {
                "$group": {
                    "_id": {
                        "$dateToString": {
                            "format": "%Y-%m-%d",
                            "date": "$fecha_transaccion"
                        }
                    },
                    "total_dia": {"$sum": "$monto"},
                    "cantidad_transacciones": {"$sum": 1},
                    "fecha": {"$first": "$fecha_transaccion"}
                }
            },
            {
                "$sort": {"_id": 1}  # Ordenar por fecha ascendente
            },
            {
                "$project": {
                    "fecha": "$_id",
                    "total_dia": 1,
                    "cantidad_transacciones": 1,
                    "_id": 0
                }
            }
        ]
        
        cursor = collection.aggregate(pipeline)
        resultados = await cursor.to_list(length=None)
        
        # Formatear los resultados
        datos_por_dia = []
        for resultado in resultados:
            datos_por_dia.append({
                "fecha": resultado["fecha"],
                "total_dia": resultado["total_dia"],
                "total_dia_formateado": f"${resultado['total_dia']:,.2f}",
                "cantidad_transacciones": resultado["cantidad_transacciones"]
            })
        
        return {
            "success": True,
            "datos_por_dia": datos_por_dia,
            "fecha_inicio": fecha_inicio,
            "fecha_fin": fecha_fin,
            "total_dias": len(datos_por_dia)
        }
        
    except ValueError as e:
        raise HTTPException(status_code=400, detail="Formato de fecha inválido. Use YYYY-MM-DD")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error al obtener datos por día: {str(e)}")

@router.get("/rutas/popularidad")
async def obtener_popularidad_rutas(
    fecha_inicio: Optional[str] = Query(None, description="Fecha de inicio en formato YYYY-MM-DD"),
    fecha_fin: Optional[str] = Query(None, description="Fecha de fin en formato YYYY-MM-DD"),
    limite: int = Query(5, description="Número máximo de rutas a retornar")
):
    """
    Obtiene la popularidad de las rutas basado en la cantidad de transacciones
    """
    try:
        collection = get_transacciones_collection()
        
        # Construir filtro de fecha si se proporciona
        filtro_fecha = {}
        if fecha_inicio and fecha_fin:
            inicio = datetime.strptime(fecha_inicio, "%Y-%m-%d")
            fin = datetime.strptime(fecha_fin, "%Y-%m-%d") + timedelta(days=1) - timedelta(seconds=1)
            
            if inicio > fin:
                raise HTTPException(status_code=400, detail="La fecha de inicio no puede ser mayor a la fecha de fin")
            
            filtro_fecha = {
                "fecha_transaccion": {
                    "$gte": inicio,
                    "$lte": fin
                }
            }
        
        pipeline = [
            {
                "$match": {
                    "tipo": "pago_viaje",
                    "nombre_ruta": {"$exists": True, "$ne": None},
                    **filtro_fecha
                }
            },
            {
                "$group": {
                    "_id": "$nombre_ruta",
                    "cantidad_viajes": {"$sum": 1},
                    "total_recaudado": {"$sum": "$monto"}
                }
            },
            {
                "$sort": {"cantidad_viajes": -1}  # Ordenar por popularidad descendente
            },
            {
                "$limit": limite  # Usar el valor numérico directamente, no Query(5)
            },
            {
                "$project": {
                    "ruta": "$_id",
                    "cantidad_viajes": 1,
                    "total_recaudado": 1,
                    "_id": 0
                }
            }
        ]
        
        cursor = collection.aggregate(pipeline)
        rutas_populares = await cursor.to_list(length=None)
        
        # Calcular totales y porcentajes
        total_viajes = sum(ruta["cantidad_viajes"] for ruta in rutas_populares)
        
        # Agregar porcentajes y formato de dinero
        for ruta in rutas_populares:
            if total_viajes > 0:
                ruta["porcentaje"] = round((ruta["cantidad_viajes"] / total_viajes) * 100, 2)
            else:
                ruta["porcentaje"] = 0
            ruta["total_recaudado_formateado"] = f"${ruta['total_recaudado']:,.2f}"
        
        return {
            "success": True,
            "rutas_populares": rutas_populares,
            "total_viajes": total_viajes,
            "fecha_inicio": fecha_inicio,
            "fecha_fin": fecha_fin,
            "limite": limite
        }
        
    except ValueError as e:
        raise HTTPException(status_code=400, detail="Formato de fecha inválido. Use YYYY-MM-DD")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error al obtener popularidad de rutas: {str(e)}")

@router.get("/estadisticas/completas")
async def obtener_estadisticas_completas(
    fecha_inicio: str = Query(..., description="Fecha de inicio en formato YYYY-MM-DD"),
    fecha_fin: str = Query(..., description="Fecha de fin en formato YYYY-MM-DD")
):
    """
    Obtiene todas las estadísticas en una sola llamada
    """
    try:
        collection = get_transacciones_collection()
        notificaciones_collection = get_notificaciones_collection()
        ruta_asignada_collection = get_ruta_asignada_collection()
        
        # Convertir strings a datetime
        inicio = datetime.strptime(fecha_inicio, "%Y-%m-%d")
        fin = datetime.strptime(fecha_fin, "%Y-%m-%d") + timedelta(days=1) - timedelta(seconds=1)
        
        # Validar que la fecha de inicio no sea mayor a la de fin
        if inicio > fin:
            raise HTTPException(status_code=400, detail="La fecha de inicio no puede ser mayor a la fecha de fin")
        
        # 1. Contar transacciones (pasajeros)
        count = await collection.count_documents({
            "tipo": "pago_viaje",
            "fecha_transaccion": {
                "$gte": inicio,
                "$lte": fin
            }
        })
        
        # 2. Sumar monto total
        pipeline_suma = [
            {
                "$match": {
                    "tipo": "pago_viaje",
                    "fecha_transaccion": {
                        "$gte": inicio,
                        "$lte": fin
                    }
                }
            },
            {
                "$group": {
                    "_id": None,
                    "total_recaudado": {"$sum": "$monto"}
                }
            }
        ]
        
        cursor_suma = collection.aggregate(pipeline_suma)
        result_suma = await cursor_suma.to_list(length=1)
        total_recaudado = result_suma[0]["total_recaudado"] if result_suma else 0
        
        # 3. Contar incidentes
        count_incidentes = await notificaciones_collection.count_documents({
            "tipo": "emergencia",
            "timestamp": {
                "$gte": inicio,
                "$lte": fin
            }
        })
        
        # 4. Calcular eficiencia
        eficiencia = 0
        tiempo_promedio = 0
        try:
            pipeline_eficiencia = [
                {
                    "$match": {
                        "fecha_asignacion": {
                            "$gte": inicio,
                            "$lte": fin
                        },
                        "tiempo_vuelta": {"$exists": True, "$ne": []}
                    }
                },
                {
                    "$unwind": "$tiempo_vuelta"
                },
                {
                    "$group": {
                        "_id": None,
                        "tiempo_promedio": {"$avg": "$tiempo_vuelta"}
                    }
                }
            ]
            
            cursor_eficiencia = ruta_asignada_collection.aggregate(pipeline_eficiencia)
            result_eficiencia = await cursor_eficiencia.to_list(length=1)
            
            if result_eficiencia:
                tiempo_promedio = result_eficiencia[0]["tiempo_promedio"]
                tiempo_ideal = 40
                eficiencia = max(0, min(100, (tiempo_ideal / tiempo_promedio) * 100)) if tiempo_promedio > 0 else 0
                eficiencia = round(eficiencia, 1)
                tiempo_promedio = round(tiempo_promedio, 2)
                
        except Exception as e:
            print(f"Advertencia: Error al calcular eficiencia: {str(e)}")
        
        # 5. Obtener datos por día (código existente)
        pipeline_dias = [
            {
                "$match": {
                    "tipo": "pago_viaje",
                    "fecha_transaccion": {
                        "$gte": inicio,
                        "$lte": fin
                    }
                }
            },
            {
                "$group": {
                    "_id": {
                        "$dateToString": {
                            "format": "%Y-%m-%d",
                            "date": "$fecha_transaccion"
                        }
                    },
                    "total_dia": {"$sum": "$monto"},
                    "cantidad_transacciones": {"$sum": 1},
                    "fecha": {"$first": "$fecha_transaccion"}
                }
            },
            {
                "$sort": {"_id": 1}
            },
            {
                "$project": {
                    "fecha": "$_id",
                    "total_dia": 1,
                    "cantidad_transacciones": 1,
                    "_id": 0
                }
            }
        ]
        
        cursor_dias = collection.aggregate(pipeline_dias)
        datos_por_dia = await cursor_dias.to_list(length=None)
        
        # Formatear datos por día
        for dia in datos_por_dia:
            dia["total_dia_formateado"] = f"${dia['total_dia']:,.2f}"
        
        # 6. Obtener rutas populares (código existente)
        rutas_populares = []
        try:
            pipeline_rutas = [
                {
                    "$match": {
                        "tipo": "pago_viaje",
                        "nombre_ruta": {"$exists": True, "$ne": None},
                        "fecha_transaccion": {
                            "$gte": inicio,
                            "$lte": fin
                        }
                    }
                },
                {
                    "$group": {
                        "_id": "$nombre_ruta",
                        "cantidad_viajes": {"$sum": 1},
                        "total_recaudado": {"$sum": "$monto"}
                    }
                },
                {
                    "$sort": {"cantidad_viajes": -1}
                },
                {
                    "$limit": 5
                },
                {
                    "$project": {
                        "ruta": "$_id",
                        "cantidad_viajes": 1,
                        "total_recaudado": 1,
                        "_id": 0
                    }
                }
            ]
            
            cursor_rutas = collection.aggregate(pipeline_rutas)
            rutas_populares = await cursor_rutas.to_list(length=None)
            
            # Calcular porcentajes para rutas
            total_viajes_rutas = sum(ruta["cantidad_viajes"] for ruta in rutas_populares)
            for ruta in rutas_populares:
                if total_viajes_rutas > 0:
                    ruta["porcentaje"] = round((ruta["cantidad_viajes"] / total_viajes_rutas) * 100, 2)
                else:
                    ruta["porcentaje"] = 0
                ruta["total_recaudado_formateado"] = f"${ruta['total_recaudado']:,.2f}"
                
        except Exception as e:
            print(f"Advertencia: Error al obtener rutas populares: {str(e)}")
        
        return {
            "success": True,
            "estadisticas": {
                "cantidad_transacciones": count,
                "total_recaudado": total_recaudado,
                "total_recaudado_formateado": f"${total_recaudado:,.2f}",
                "total_incidentes": count_incidentes,
                "eficiencia": eficiencia,
                "tiempo_promedio": tiempo_promedio,
                "datos_por_dia": datos_por_dia,
                "rutas_populares": rutas_populares
            },
            "fecha_inicio": fecha_inicio,
            "fecha_fin": fecha_fin
        }
        
    except ValueError as e:
        raise HTTPException(status_code=400, detail="Formato de fecha inválido. Use YYYY-MM-DD")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error al obtener estadísticas completas: {str(e)}")
    
@router.get("/incidentes/count")
async def contar_incidentes_por_periodo(
    fecha_inicio: str = Query(..., description="Fecha de inicio en formato YYYY-MM-DD"),
    fecha_fin: str = Query(..., description="Fecha de fin en formato YYYY-MM-DD")
):
    """
    Cuenta incidentes (notificaciones de emergencia) en un periodo específico
    """
    try:
        collection = get_notificaciones_collection()
        
        # Convertir strings a datetime
        inicio = datetime.strptime(fecha_inicio, "%Y-%m-%d")
        fin = datetime.strptime(fecha_fin, "%Y-%m-%d") + timedelta(days=1) - timedelta(seconds=1)
        
        # Validar que la fecha de inicio no sea mayor a la de fin
        if inicio > fin:
            raise HTTPException(status_code=400, detail="La fecha de inicio no puede ser mayor a la fecha de fin")
        
        count = await collection.count_documents({
            "tipo": "emergencia",
            "timestamp": {
                "$gte": inicio,
                "$lte": fin
            }
        })
        
        return {
            "success": True,
            "count": count,
            "fecha_inicio": fecha_inicio,
            "fecha_fin": fecha_fin,
            "tipo": "emergencia"
        }
        
    except ValueError as e:
        raise HTTPException(status_code=400, detail="Formato de fecha inválido. Use YYYY-MM-DD")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error al contar incidentes: {str(e)}")

@router.get("/eficiencia/promedio")
async def calcular_eficiencia_rutas(
    fecha_inicio: str = Query(..., description="Fecha de inicio en formato YYYY-MM-DD"),
    fecha_fin: str = Query(..., description="Fecha de fin en formato YYYY-MM-DD")
):
    """
    Calcula la eficiencia de rutas basado en el tiempo promedio de vuelta
    """
    try:
        collection = get_ruta_asignada_collection()
        
        # Convertir strings a datetime
        inicio = datetime.strptime(fecha_inicio, "%Y-%m-%d")
        fin = datetime.strptime(fecha_fin, "%Y-%m-%d") + timedelta(days=1) - timedelta(seconds=1)
        
        # Validar que la fecha de inicio no sea mayor a la de fin
        if inicio > fin:
            raise HTTPException(status_code=400, detail="La fecha de inicio no puede ser mayor a la fecha de fin")
        
        # Pipeline para calcular estadísticas de tiempo de vuelta
        pipeline = [
            {
                "$match": {
                    "fecha_asignacion": {
                        "$gte": inicio,
                        "$lte": fin
                    },
                    "tiempo_vuelta": {"$exists": True, "$ne": []}
                }
            },
            {
                "$unwind": "$tiempo_vuelta"
            },
            {
                "$group": {
                    "_id": None,
                    "total_tiempos": {"$sum": 1},
                    "suma_tiempos": {"$sum": "$tiempo_vuelta"},
                    "tiempo_promedio": {"$avg": "$tiempo_vuelta"}
                }
            }
        ]
        
        cursor = collection.aggregate(pipeline)
        result = await cursor.to_list(length=1)
        
        if result:
            tiempo_promedio = result[0]["tiempo_promedio"]
            total_tiempos = result[0]["total_tiempos"]
            suma_tiempos = result[0]["suma_tiempos"]
            
            # Calcular eficiencia (tiempo ideal: 40 minutos = 100%)
            tiempo_ideal = 40
            eficiencia = max(0, min(100, (tiempo_ideal / tiempo_promedio) * 100)) if tiempo_promedio > 0 else 0
            
            return {
                "success": True,
                "eficiencia": round(eficiencia, 1),
                "tiempo_promedio": round(tiempo_promedio, 2),
                "total_vueltas_analizadas": total_tiempos,
                "tiempo_ideal": tiempo_ideal,
                "fecha_inicio": fecha_inicio,
                "fecha_fin": fecha_fin
            }
        else:
            return {
                "success": True,
                "eficiencia": 0,
                "tiempo_promedio": 0,
                "total_vueltas_analizadas": 0,
                "tiempo_ideal": 40,
                "fecha_inicio": fecha_inicio,
                "fecha_fin": fecha_fin,
                "mensaje": "No hay datos de tiempos de vuelta en el periodo seleccionado"
            }
        
    except ValueError as e:
        raise HTTPException(status_code=400, detail="Formato de fecha inválido. Use YYYY-MM-DD")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error al calcular eficiencia: {str(e)}")

@router.get("/incidentes/detalle")
async def obtener_incidentes_detalle(
    fecha_inicio: str = Query(..., description="Fecha de inicio en formato YYYY-MM-DD"),
    fecha_fin: str = Query(..., description="Fecha de fin en formato YYYY-MM-DD")
):
    """
    Obtiene el detalle de incidentes (notificaciones de emergencia) en un periodo
    """
    try:
        collection = get_notificaciones_collection()
        
        # Convertir strings a datetime
        inicio = datetime.strptime(fecha_inicio, "%Y-%m-%d")
        fin = datetime.strptime(fecha_fin, "%Y-%m-%d") + timedelta(days=1) - timedelta(seconds=1)
        
        # Validar que la fecha de inicio no sea mayor a la de fin
        if inicio > fin:
            raise HTTPException(status_code=400, detail="La fecha de inicio no puede ser mayor a la fecha de fin")
        
        incidentes = await collection.find({
            "tipo": "emergencia",
            "timestamp": {
                "$gte": inicio,
                "$lte": fin
            }
        }).sort("timestamp", -1).to_list(length=100)
        
        # Formatear los resultados
        incidentes_formateados = []
        for incidente in incidentes:
            incidente_formateado = {
                "id_notificacion": incidente.get("id_notificacion"),
                "titulo": incidente.get("titulo"),
                "mensaje": incidente.get("mensaje"),
                "ruta_afectada": incidente.get("ruta_afectada"),
                "timestamp": incidente.get("timestamp"),
                "leida": incidente.get("leida", False)
            }
            
            # Formatear timestamp si es datetime
            if isinstance(incidente_formateado["timestamp"], datetime):
                incidente_formateado["timestamp_formateado"] = incidente_formateado["timestamp"].strftime("%Y-%m-%d %H:%M")
            
            incidentes_formateados.append(incidente_formateado)
        
        return {
            "success": True,
            "incidentes": incidentes_formateados,
            "total_incidentes": len(incidentes_formateados),
            "fecha_inicio": fecha_inicio,
            "fecha_fin": fecha_fin
        }
        
    except ValueError as e:
        raise HTTPException(status_code=400, detail="Formato de fecha inválido. Use YYYY-MM-DD")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error al obtener detalle de incidentes: {str(e)}")
    
@router.get("/exportar/excel")
async def exportar_reportes_excel(
    fecha_inicio: str = Query(..., description="Fecha de inicio en formato YYYY-MM-DD"),
    fecha_fin: str = Query(..., description="Fecha de fin en formato YYYY-MM-DD")
):
    """
    Exporta todos los reportes a Excel con gráficas incluidas
    """
    try:
        # Convertir strings a datetime
        inicio = datetime.strptime(fecha_inicio, "%Y-%m-%d")
        fin = datetime.strptime(fecha_fin, "%Y-%m-%d") + timedelta(days=1) - timedelta(seconds=1)
        
        # Validar que la fecha de inicio no sea mayor a la de fin
        if inicio > fin:
            raise HTTPException(status_code=400, detail="La fecha de inicio no puede ser mayor a la fecha de fin")
        
        # Obtener todas las estadísticas
        transacciones_collection = get_transacciones_collection()
        notificaciones_collection = get_notificaciones_collection()
        ruta_asignada_collection = get_ruta_asignada_collection()
        
        # 1. Estadísticas generales
        pipeline_estadisticas = [
            {
                "$match": {
                    "tipo": "pago_viaje",
                    "fecha_transaccion": {
                        "$gte": inicio,
                        "$lte": fin
                    }
                }
            },
            {
                "$group": {
                    "_id": None,
                    "total_recaudado": {"$sum": "$monto"},
                    "cantidad_transacciones": {"$sum": 1}
                }
            }
        ]
        
        cursor_estadisticas = transacciones_collection.aggregate(pipeline_estadisticas)
        result_estadisticas = await cursor_estadisticas.to_list(length=1)
        
        total_recaudado = result_estadisticas[0]["total_recaudado"] if result_estadisticas else 0
        cantidad_transacciones = result_estadisticas[0]["cantidad_transacciones"] if result_estadisticas else 0
        
        # 2. Contar incidentes
        count_incidentes = await notificaciones_collection.count_documents({
            "tipo": "emergencia",
            "timestamp": {
                "$gte": inicio,
                "$lte": fin
            }
        })
        
        # 3. Calcular eficiencia
        pipeline_eficiencia = [
            {
                "$match": {
                    "fecha_asignacion": {
                        "$gte": inicio,
                        "$lte": fin
                    },
                    "tiempo_vuelta": {"$exists": True, "$ne": []}
                }
            },
            {
                "$unwind": "$tiempo_vuelta"
            },
            {
                "$group": {
                    "_id": None,
                    "tiempo_promedio": {"$avg": "$tiempo_vuelta"}
                }
            }
        ]
        
        cursor_eficiencia = ruta_asignada_collection.aggregate(pipeline_eficiencia)
        result_eficiencia = await cursor_eficiencia.to_list(length=1)
        
        eficiencia = 0
        tiempo_promedio = 0
        if result_eficiencia:
            tiempo_promedio = result_eficiencia[0]["tiempo_promedio"]
            tiempo_ideal = 40
            eficiencia = max(0, min(100, (tiempo_ideal / tiempo_promedio) * 100)) if tiempo_promedio > 0 else 0
            eficiencia = round(eficiencia, 1)
            tiempo_promedio = round(tiempo_promedio, 2)
        
        # 4. Datos por día
        pipeline_dias = [
            {
                "$match": {
                    "tipo": "pago_viaje",
                    "fecha_transaccion": {
                        "$gte": inicio,
                        "$lte": fin
                    }
                }
            },
            {
                "$group": {
                    "_id": {
                        "$dateToString": {
                            "format": "%Y-%m-%d",
                            "date": "$fecha_transaccion"
                        }
                    },
                    "total_dia": {"$sum": "$monto"},
                    "cantidad_transacciones": {"$sum": 1}
                }
            },
            {
                "$sort": {"_id": 1}
            },
            {
                "$project": {
                    "fecha": "$_id",
                    "total_dia": 1,
                    "cantidad_transacciones": 1,
                    "_id": 0
                }
            }
        ]
        
        cursor_dias = transacciones_collection.aggregate(pipeline_dias)
        datos_por_dia = await cursor_dias.to_list(length=None)
        
        # 5. Rutas populares
        pipeline_rutas = [
            {
                "$match": {
                    "tipo": "pago_viaje",
                    "nombre_ruta": {"$exists": True, "$ne": None},
                    "fecha_transaccion": {
                        "$gte": inicio,
                        "$lte": fin
                    }
                }
            },
            {
                "$group": {
                    "_id": "$nombre_ruta",
                    "cantidad_viajes": {"$sum": 1},
                    "total_recaudado": {"$sum": "$monto"}
                }
            },
            {
                "$sort": {"cantidad_viajes": -1}
            },
            {
                "$limit": 10
            },
            {
                "$project": {
                    "ruta": "$_id",
                    "cantidad_viajes": 1,
                    "total_recaudado": 1,
                    "_id": 0
                }
            }
        ]
        
        cursor_rutas = transacciones_collection.aggregate(pipeline_rutas)
        rutas_populares = await cursor_rutas.to_list(length=None)
        
        # 6. Incidentes detallados
        incidentes_detalle = await notificaciones_collection.find({
            "tipo": "emergencia",
            "timestamp": {
                "$gte": inicio,
                "$lte": fin
            }
        }).sort("timestamp", -1).to_list(length=100)
        
        # Crear Excel con múltiples hojas y gráficas
        output = io.BytesIO()
        
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            workbook = writer.book
            
            # Estilos para el Excel
            title_font = Font(size=14, bold=True, color="FFFFFF")
            header_font = Font(bold=True, color="FFFFFF")
            title_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            
            # Hoja 1: Resumen Ejecutivo
            resumen_data = {
                'Métrica': [
                    'Fecha Inicio',
                    'Fecha Fin',
                    'Total Recaudado',
                    'Cantidad de Transacciones',
                    'Pasajeros Totales',
                    'Incidentes Reportados',
                    'Eficiencia de Rutas',
                    'Tiempo Promedio de Vuelta (min)'
                ],
                'Valor': [
                    fecha_inicio,
                    fecha_fin,
                    f"${total_recaudado:,.2f}",
                    cantidad_transacciones,
                    cantidad_transacciones,
                    count_incidentes,
                    f"{eficiencia}%",
                    f"{tiempo_promedio} min"
                ]
            }
            df_resumen = pd.DataFrame(resumen_data)
            df_resumen.to_excel(writer, sheet_name='Resumen Ejecutivo', index=False, startrow=1)
            
            worksheet_resumen = writer.sheets['Resumen Ejecutivo']
            
            # Título de la hoja
            worksheet_resumen.merge_cells('A1:B1')
            title_cell = worksheet_resumen['A1']
            title_cell.value = f'REPORTE EJECUTIVO TUCSA - {fecha_inicio} a {fecha_fin}'
            title_cell.font = title_font
            title_cell.fill = title_fill
            title_cell.alignment = Alignment(horizontal='center')
            
            # Formatear encabezados
            for cell in worksheet_resumen['A2:B2'][0]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center')
            
            # Ajustar anchos de columna
            worksheet_resumen.column_dimensions['A'].width = 30
            worksheet_resumen.column_dimensions['B'].width = 25
            
            # Hoja 2: Ingresos por Día con Gráfica
            if datos_por_dia:
                datos_dias_formateados = []
                for dia in datos_por_dia:
                    fecha_obj = datetime.strptime(dia['fecha'], "%Y-%m-%d")
                    fecha_formateada = fecha_obj.strftime("%d/%m/%Y")
                    datos_dias_formateados.append({
                        'Fecha': fecha_formateada,
                        'Fecha Orden': fecha_obj,
                        'Ingresos': dia['total_dia'],
                        'Ingresos Formateado': f"${dia['total_dia']:,.2f}",
                        'Cantidad Transacciones': dia['cantidad_transacciones']
                    })
                
                # Ordenar por fecha
                datos_dias_formateados.sort(key=lambda x: x['Fecha Orden'])
                
                df_dias = pd.DataFrame(datos_dias_formateados)
                df_dias = df_dias[['Fecha', 'Ingresos', 'Ingresos Formateado', 'Cantidad Transacciones']]
                df_dias.to_excel(writer, sheet_name='Ingresos por Día', index=False, startrow=1)
                
                worksheet_dias = writer.sheets['Ingresos por Día']
                
                # Título
                worksheet_dias.merge_cells('A1:D1')
                title_cell = worksheet_dias['A1']
                title_cell.value = 'INGRESOS POR DÍA'
                title_cell.font = title_font
                title_cell.fill = title_fill
                title_cell.alignment = Alignment(horizontal='center')
                
                # Formatear encabezados
                for cell in worksheet_dias['A2:D2'][0]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal='center')
                
                # Crear gráfica de barras
                chart1 = BarChart()
                chart1.type = "col"
                chart1.style = 10
                chart1.title = "Ingresos por Día"
                chart1.y_axis.title = 'Ingresos ($)'
                chart1.x_axis.title = 'Fecha'
                
                data = Reference(worksheet_dias, min_col=2, min_row=2, max_row=len(datos_dias_formateados) + 2, max_col=2)
                categories = Reference(worksheet_dias, min_col=1, min_row=3, max_row=len(datos_dias_formateados) + 2)
                
                chart1.add_data(data, titles_from_data=True)
                chart1.set_categories(categories)
                chart1.shape = 4
                
                # Posicionar gráfica
                worksheet_dias.add_chart(chart1, "F2")
                
                # Ajustar anchos
                worksheet_dias.column_dimensions['A'].width = 15
                worksheet_dias.column_dimensions['B'].width = 15
                worksheet_dias.column_dimensions['C'].width = 20
                worksheet_dias.column_dimensions['D'].width = 20
                
            else:
                df_dias = pd.DataFrame({'Mensaje': ['No hay datos para el periodo seleccionado']})
                df_dias.to_excel(writer, sheet_name='Ingresos por Día', index=False)
            
            # Hoja 3: Rutas Populares con Gráfica
            if rutas_populares:
                rutas_formateadas = []
                for ruta in rutas_populares:
                    rutas_formateadas.append({
                        'Ruta': ruta['ruta'],
                        'Viajes Realizados': ruta['cantidad_viajes'],
                        'Total Recaudado': ruta['total_recaudado'],
                        'Total Recaudado Formateado': f"${ruta['total_recaudado']:,.2f}"
                    })
                df_rutas = pd.DataFrame(rutas_formateadas)
                df_rutas.to_excel(writer, sheet_name='Rutas Populares', index=False, startrow=1)
                
                worksheet_rutas = writer.sheets['Rutas Populares']
                
                # Título
                worksheet_rutas.merge_cells('A1:D1')
                title_cell = worksheet_rutas['A1']
                title_cell.value = 'RUTAS MÁS POPULARES'
                title_cell.font = title_font
                title_cell.fill = title_fill
                title_cell.alignment = Alignment(horizontal='center')
                
                # Formatear encabezados
                for cell in worksheet_rutas['A2:D2'][0]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal='center')
                
                # Crear gráfica de pastel
                pie_chart = PieChart()
                pie_chart.title = "Distribución de Viajes por Ruta"
                
                labels = Reference(worksheet_rutas, min_col=1, min_row=3, max_row=len(rutas_populares) + 2)
                data = Reference(worksheet_rutas, min_col=2, min_row=2, max_row=len(rutas_populares) + 2)
                
                pie_chart.add_data(data, titles_from_data=True)
                pie_chart.set_categories(labels)
                pie_chart.style = 10
                
                # Posicionar gráfica
                worksheet_rutas.add_chart(pie_chart, "F2")
                
                # Crear gráfica de barras para ingresos por ruta
                chart2 = BarChart()
                chart2.type = "col"
                chart2.style = 11
                chart2.title = "Ingresos por Ruta"
                chart2.y_axis.title = 'Ingresos ($)'
                
                data_ingresos = Reference(worksheet_rutas, min_col=3, min_row=2, max_row=len(rutas_populares) + 2, max_col=3)
                categories_rutas = Reference(worksheet_rutas, min_col=1, min_row=3, max_row=len(rutas_populares) + 2)
                
                chart2.add_data(data_ingresos, titles_from_data=True)
                chart2.set_categories(categories_rutas)
                
                # Posicionar segunda gráfica
                worksheet_rutas.add_chart(chart2, "F18")
                
                # Ajustar anchos
                worksheet_rutas.column_dimensions['A'].width = 25
                worksheet_rutas.column_dimensions['B'].width = 18
                worksheet_rutas.column_dimensions['C'].width = 18
                worksheet_rutas.column_dimensions['D'].width = 20
                
            else:
                df_rutas = pd.DataFrame({'Mensaje': ['No hay datos de rutas para el periodo seleccionado']})
                df_rutas.to_excel(writer, sheet_name='Rutas Populares', index=False)
            
            # Hoja 4: Incidentes Detallados
            if incidentes_detalle:
                incidentes_formateados = []
                for incidente in incidentes_detalle:
                    timestamp = incidente.get('timestamp')
                    if isinstance(timestamp, datetime):
                        fecha_formateada = timestamp.strftime("%Y-%m-%d %H:%M")
                    else:
                        fecha_formateada = str(timestamp)
                    
                    incidentes_formateados.append({
                        'ID': incidente.get('id_notificacion', ''),
                        'Título': incidente.get('titulo', ''),
                        'Mensaje': incidente.get('mensaje', ''),
                        'Ruta Afectada': incidente.get('ruta_afectada', ''),
                        'Fecha/Hora': fecha_formateada,
                        'Leído': 'Sí' if incidente.get('leida', False) else 'No'
                    })
                df_incidentes = pd.DataFrame(incidentes_formateados)
                df_incidentes.to_excel(writer, sheet_name='Incidentes', index=False, startrow=1)
                
                worksheet_incidentes = writer.sheets['Incidentes']
                
                # Título
                worksheet_incidentes.merge_cells('A1:F1')
                title_cell = worksheet_incidentes['A1']
                title_cell.value = 'INCIDENTES REPORTADOS'
                title_cell.font = title_font
                title_cell.fill = title_fill
                title_cell.alignment = Alignment(horizontal='center')
                
                # Formatear encabezados
                for cell in worksheet_incidentes['A2:F2'][0]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal='center')
                
                # Ajustar anchos
                worksheet_incidentes.column_dimensions['A'].width = 10
                worksheet_incidentes.column_dimensions['B'].width = 25
                worksheet_incidentes.column_dimensions['C'].width = 50
                worksheet_incidentes.column_dimensions['D'].width = 15
                worksheet_incidentes.column_dimensions['E'].width = 20
                worksheet_incidentes.column_dimensions['F'].width = 10
                
            else:
                df_incidentes = pd.DataFrame({'Mensaje': ['No hay incidentes reportados en el periodo seleccionado']})
                df_incidentes.to_excel(writer, sheet_name='Incidentes', index=False)
            
        
        output.seek(0)
        
        # Generar nombre del archivo
        nombre_archivo = f"reporte_tucsa_{fecha_inicio}_a_{fecha_fin}.xlsx"
        
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={nombre_archivo}"}
        )
        
    except ValueError as e:
        raise HTTPException(status_code=400, detail="Formato de fecha inválido. Use YYYY-MM-DD")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error al exportar reporte Excel: {str(e)}")